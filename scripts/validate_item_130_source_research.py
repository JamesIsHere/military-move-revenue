#!/usr/bin/env python3
"""Validate the source-only Item 130 gap research checkpoint."""

from __future__ import annotations

import copy
import json
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RESEARCH_PATH = ROOT / "docs" / "decisions" / "0005-item-130-source-research-2026-08-07.json"
REGISTRY_PATH = ROOT / "rules" / "registry" / "registry.json"
EXPECTED_GAPS = {
    "GAP-130-LAWNMOWER-ROW": "OPEN_NO_MAPPING_AUTHORITY",
    "GAP-130E-SUBTYPE-ROWS": "OPEN_NO_SUBTYPE_CROSSWALK",
    "GAP-130F-BOTO-BOUNDARY": "OPEN_CONFLICT_NARROWED",
    "GAP-130-COMBINED-VS-OD": "OPEN_NO_QUANTITY_OR_ROW_SELECTION_CONTRACT",
}
EXPECTED_CLAIMS = {f"CLM-{number:04d}" for number in range(47, 56)} | {"CLM-0037"}


class ValidationError(Exception):
    pass


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ValidationError(message)


def load_json(path: Path) -> dict:
    def reject_float(value: str) -> None:
        raise ValidationError(f"{path.name} contains non-exact JSON number {value}")

    with path.open(encoding="utf-8") as handle:
        value = json.load(handle, parse_float=reject_float)
    require(isinstance(value, dict), f"{path.name} must contain an object")
    return value


def item_rows() -> list[list[object]]:
    archive = ROOT / "sources" / "raw" / "2026" / "item-code-listing-2022-08-12.zip"
    with ZipFile(archive) as outer:
        member = next((name for name in outer.namelist() if name.lower().endswith(".xlsx")), None)
        require(member is not None, "item-code archive lacks xlsx member")
        workbook = load_workbook(BytesIO(outer.read(member)), read_only=True, data_only=False)
    require("DOM_400NG" in workbook.sheetnames, "item-code workbook lacks DOM_400NG")
    sheet = workbook["DOM_400NG"]
    return [[sheet.cell(row, column).value for column in range(1, 18)] for row in range(53, 119)]


def validate(research: dict, registry: dict) -> None:
    require(research.get("schema_version") == "item-130-source-research.v1", "research schema mismatch")
    require(research.get("research_id") == "ITEM-130-SOURCE-PASS-2026-08-07", "research id mismatch")
    require(research.get("decision_number") == "0005", "decision link mismatch")
    require(research.get("conducted_on") == "2026-08-07", "research date mismatch")
    require(research.get("status") == "COMPLETED_NO_GAP_CLOSED", "research status overstates closure")
    require(research.get("scope") == "DOMESTIC_400NG_ITEM_130_FOUR_RECORDED_SOURCE_GAPS_ONLY", "research scope mismatch")

    versions = {record["id"]: record for record in registry["source_versions"]}
    claims = {record["id"]: record for record in registry["source_claims"]}
    basis = research.get("source_basis")
    require(isinstance(basis, list) and len(basis) == 3, "source basis must contain three archived records")
    cited_claims: set[str] = set()
    provenance_ids: set[str] = set()
    for source in basis:
        provenance_id = source.get("provenance_id")
        require(isinstance(provenance_id, str) and provenance_id.startswith("P-130R-") and provenance_id not in provenance_ids, "source provenance id invalid or duplicated")
        provenance_ids.add(provenance_id)
        version_id = source.get("source_version_id")
        require(version_id in versions and versions[version_id]["source_id"] == source.get("source_id"), f"{provenance_id} source/version mismatch")
        for field in ("document_version", "effective_period", "locator", "retrieval_date", "interpretation_status"):
            require(isinstance(source.get(field), str) and source[field], f"{provenance_id} lacks {field}")
        source_claims = source.get("source_claim_ids")
        require(isinstance(source_claims, list) and source_claims, f"{provenance_id} lacks source claims")
        require(all(claim_id in claims for claim_id in source_claims), f"{provenance_id} cites an unknown claim")
        cited_claims.update(source_claims)

    require(cited_claims == EXPECTED_CLAIMS, "research source-claim set changed")
    findings = research.get("findings")
    require(isinstance(findings, list) and len(findings) == 4, "research must contain four findings")
    require({row.get("gap_id"): row.get("status") for row in findings} == EXPECTED_GAPS, "gap status set changed")
    for finding in findings:
        require(finding.get("finding") and finding.get("prohibited_inference"), f"{finding.get('gap_id')} lacks finding or stop gate")
        claim_ids = finding.get("source_claim_ids")
        require(isinstance(claim_ids, list) and len(claim_ids) >= 2, f"{finding.get('gap_id')} lacks opposing source claims")
        require(all(claim_id in cited_claims for claim_id in claim_ids), f"{finding.get('gap_id')} cites an unregistered claim")

    financial = research.get("financial_contract")
    require(isinstance(financial, dict) and financial.get("status") == "PROHIBITED", "financial contract is not prohibited")
    for field in ("rate_version_date_fact", "billing_item_contract", "quantity_for_billing", "rate", "expected_amount", "audit_adapter"):
        require(financial.get(field) is None, f"financial field {field} must remain null")
    external = research.get("external_action")
    require(external.get("rates_team_request_status") == "NOT_SENT_APPROVAL_REQUIRED", "external request status changed")
    require(len(external.get("required_clarification", [])) == 5, "clarification question set changed")

    rows = item_rows()
    require(len(rows) == 66, "Item 130 row count changed")
    require(not any("lawn" in str(row[5]).lower() for row in rows), "lawnmower gap no longer matches archive")
    require([str(row[5]).strip() for row in rows if row[1] == "130E"] == ["Bulky Article: Boats > 14 Ft.", "Bulky Article: Boats > 14 Ft."], "130E descriptions changed")
    rows_130f = [row for row in rows if row[1] == "130F"]
    require([(row[4], str(row[15]).strip()) for row in rows_130f] == [("O", "Origin PPSO"), ("D", "Destin PPSO")], "130F origin/destination representations changed")
    require(all(rows[index][4] == ("O" if index % 2 == 0 else "D") for index in range(len(rows))), "Item 130 origin/destination pair structure changed")

    tariff = " ".join((ROOT / "sources" / "derived" / "2026" / "2026-400ng-final.txt").read_text(encoding="utf-8").split())
    required_tariff_text = (
        "Riding Lawnmowers (including stand-on)",
        "Boats >14 ft.., Dinghies, Sculls, Skiffs, or Row Boats",
        "Boat trailers are handled under the BOTO program",
        "One charge for each time a combined loading and unloading service is required",
        "When the TSP transporting the HHGs does not agree to move a boat (over 14’) with the HHGs",
    )
    for phrase in required_tariff_text:
        require(phrase in tariff, f"tariff extract lacks {phrase}")
    library = (ROOT / "sources" / "raw" / "2026" / "ustranscom-dp3-library-2026-08-03.html").read_text(encoding="utf-8", errors="replace")
    require("Item Code Listing (12 Aug 2022).zip" in library, "library snapshot lacks item-code link")


def main() -> int:
    try:
        research = load_json(RESEARCH_PATH)
        registry = load_json(REGISTRY_PATH)
        validate(research, registry)
        probes = [
            ("closure", lambda value: value.__setitem__("status", "RESOLVED")),
            ("gap", lambda value: value["findings"].pop()),
            ("130F status", lambda value: value["findings"][2].__setitem__("status", "RESOLVED")),
            ("money", lambda value: value["financial_contract"].__setitem__("expected_amount", "297.78")),
            ("request", lambda value: value["external_action"].__setitem__("rates_team_request_status", "SENT")),
            ("claim", lambda value: value["source_basis"][0]["source_claim_ids"].pop()),
        ]
        for label, mutate in probes:
            changed = copy.deepcopy(research)
            mutate(changed)
            try:
                validate(changed, registry)
            except ValidationError:
                print(f"PASS Item 130 source research tamper rejected: {label}")
                continue
            raise ValidationError(f"Item 130 source research tamper accepted: {label}")
        print("PASS Item 130 source research: 3 archived sources, 10 registered claims, 4 open gaps, and 6 tamper probes")
        return 0
    except (OSError, StopIteration, json.JSONDecodeError, KeyError, TypeError, ValidationError) as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
