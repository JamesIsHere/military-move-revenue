"""Create a provenance-preserving structural extract of archived public workbooks.

This is a read-only fallback for sessions where @oai/artifact-tool is unavailable.
It never saves a workbook. ZIP members are read into memory and source artifacts
are verified by SHA-256 before inspection.
"""

from __future__ import annotations

import hashlib
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "sources" / "raw" / "2026"
OUTPUT = ROOT / "sources" / "derived" / "2026" / "workbook-structure.json"

SOURCES = (
    {
        "source_id": "SRC-DP3-2026-RATES",
        "artifact": RAW / "400ng-baseline-rates.zip",
        "member": "400NG Baseline Rates.xlsx",
        "version_or_publication": "2026",
        "effective_period": "2026-05-15/2027-05-14",
        "retrieved_on": "2026-08-03",
        "match_terms": ("185A", "185B", "210A", "210B", "210C", "210D", "210E", "210F"),
        "configured_ranges": {
            "Base Point City": ("A1:E8",),
            "Geographical Schedule": ("A1:H8",),
            "Additional Rates": ("A52:F66",),
            "Accessorials": ("A1:Z10",),
        },
    },
    {
        "source_id": "SRC-DP3-2026-TRANSIT",
        "artifact": RAW / "2026-transit-time-tables.zip",
        "member": (
            "2026-USTC Dom-Intern Transit Time Tables- Effective Date "
            "15 May 2026 - 8 Dec 25-Final.xlsx"
        ),
        "version_or_publication": "2025-12-08",
        "effective_period": "2026-05-15/open",
        "retrieved_on": "2026-08-03",
        "match_terms": ("MILES", "Transit Times"),
        "configured_ranges": {"Appendix L-Domestic": ("A1:F33",)},
    },
    {
        "source_id": "SRC-DP3-ITEM-CODES",
        "artifact": RAW / "item-code-listing-2022-08-12.zip",
        "member": "Item Code Listing (12 Aug 2022).xlsx",
        "version_or_publication": "2022-08-12",
        "effective_period": "not stated",
        "retrieved_on": "2026-08-03",
        "match_terms": (
            "17A", "17B", "17C", "17D", "17E", "17F", "17G",
            "185A", "185B", "210A", "210B", "210C", "210D", "210E", "210F", "226A",
        ),
        "configured_ranges": {
            "DOM_400NG": ("A1:Q15", "A120:Q146", "A147:Q166"),
        },
    },
    {
        "source_id": "SRC-DP3-MILEAGE-SIT",
        "artifact": RAW / "dps-mileage-transit-time-sit-tool.xlsx",
        "member": None,
        "version_or_publication": "current at retrieval",
        "effective_period": "not stated",
        "retrieved_on": "2026-08-03",
        "match_terms": ("Weight", "Origin Zip", "Destination Zip", "Miles", "Transit", "SIT"),
        "configured_ranges": {
            "MAIN": ("C2:H13",),
            "WORK": ("A1:I5",),
            "EREF": ("A2:B5",),
            "TREF": ("A1:C10",),
            "TT": ("A1:F21",),
        },
    },
)


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest().upper()


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Inspection only: preserve the displayed lexical value and do no math.
        return str(value)
    return str(value)


def cell_record(cell: Any) -> dict[str, Any]:
    return {
        "coordinate": cell.coordinate,
        "value": json_value(cell.value),
        "data_type": cell.data_type,
        "number_format": cell.number_format,
        "style_id": cell.style_id,
    }


def inspect_range(ws: Any, range_ref: str) -> dict[str, Any]:
    rows = []
    for row in ws[range_ref]:
        populated = [cell for cell in row if cell.value is not None]
        if populated:
            rows.append(
                {
                    "row": populated[0].row,
                    "cells": [cell_record(cell) for cell in populated],
                }
            )
    return {"range": range_ref, "rows": rows}


def inspect_sheet(
    ws: Any,
    cached_ws: Any,
    match_terms: tuple[str, ...],
    configured_ranges: tuple[str, ...],
) -> dict[str, Any]:
    first_rows: list[dict[str, Any]] = []
    last_rows: list[dict[str, Any]] = []
    formula_samples: list[dict[str, Any]] = []
    formula_count = 0
    formula_cached_error_count = 0
    formula_cached_error_samples: list[dict[str, Any]] = []
    nonempty_cells = 0
    min_row = None
    min_col = None
    max_row = 0
    max_col = 0
    targeted_row_matches: list[dict[str, Any]] = []

    for row in ws.iter_rows():
        populated = [cell for cell in row if cell.value is not None]
        if not populated:
            continue

        row_index = populated[0].row
        nonempty_cells += len(populated)
        min_row = row_index if min_row is None else min(min_row, row_index)
        min_col_value = min(cell.column for cell in populated)
        min_col = min_col_value if min_col is None else min(min_col, min_col_value)
        max_row = max(max_row, row_index)
        max_col = max(max_col, max(cell.column for cell in populated))

        row_record = {
            "row": row_index,
            "cells": [cell_record(cell) for cell in populated[:24]],
        }
        if len(first_rows) < 12:
            first_rows.append(row_record)
        last_rows.append(row_record)
        if len(last_rows) > 3:
            last_rows.pop(0)

        row_text = " | ".join(str(cell.value) for cell in populated)
        if (
            len(targeted_row_matches) < 120
            and any(term.casefold() in row_text.casefold() for term in match_terms)
        ):
            targeted_row_matches.append(row_record)

        for cell in populated:
            if cell.data_type == "f":
                formula_count += 1
                cached_cell = cached_ws[cell.coordinate]
                if cached_cell.data_type == "e":
                    formula_cached_error_count += 1
                    if len(formula_cached_error_samples) < 20:
                        formula_cached_error_samples.append(cell_record(cached_cell))
                if len(formula_samples) < 20:
                    formula_record = cell_record(cell)
                    formula_record["cached_value"] = json_value(cached_cell.value)
                    formula_samples.append(formula_record)

    actual_range = None
    if min_row is not None and min_col is not None:
        actual_range = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )

    validations = []
    if ws.data_validations is not None:
        for validation in ws.data_validations.dataValidation:
            validations.append(
                {
                    "type": validation.type,
                    "sqref": str(validation.sqref),
                    "formula1": json_value(validation.formula1),
                    "formula2": json_value(validation.formula2),
                    "allow_blank": validation.allow_blank,
                }
            )

    tables = []
    for name in ws.tables:
        table = ws.tables[name]
        tables.append({"name": name, "ref": table.ref})

    return {
        "title": ws.title,
        "state": ws.sheet_state,
        "declared_dimension": ws.calculate_dimension(),
        "actual_nonempty_range": actual_range,
        "nonempty_cell_count": nonempty_cells,
        "formula_count": formula_count,
        "formula_cached_error_count": formula_cached_error_count,
        "formula_cached_error_samples": formula_cached_error_samples,
        "formula_samples": formula_samples,
        "first_nonempty_rows": first_rows,
        "last_nonempty_rows": last_rows,
        "targeted_row_matches": targeted_row_matches,
        "merged_range_count": len(ws.merged_cells.ranges),
        "merged_range_samples": [str(value) for value in list(ws.merged_cells.ranges)[:20]],
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "auto_filter_ref": ws.auto_filter.ref,
        "tables": tables,
        "data_validations": validations,
        "conditional_formatting_rule_count": len(ws.conditional_formatting),
        "configured_range_extracts": [
            inspect_range(ws, range_ref) for range_ref in configured_ranges
        ],
        "configured_range_cached_extracts": [
            inspect_range(cached_ws, range_ref) for range_ref in configured_ranges
        ],
    }


def read_workbook_bytes(source: dict[str, Any]) -> tuple[bytes, bytes]:
    artifact_bytes = source["artifact"].read_bytes()
    if source["member"] is None:
        return artifact_bytes, artifact_bytes
    with zipfile.ZipFile(BytesIO(artifact_bytes)) as archive:
        workbook_bytes = archive.read(source["member"])
    return artifact_bytes, workbook_bytes


def read_core_properties(workbook_bytes: bytes) -> dict[str, Any]:
    """Preserve core-property text exactly as stored in the XLSX package."""
    with zipfile.ZipFile(BytesIO(workbook_bytes)) as archive:
        root = ET.fromstring(archive.read("docProps/core.xml"))
    properties = {child.tag.rsplit("}", 1)[-1]: child.text for child in root}
    return {
        "title": properties.get("title"),
        "creator": properties.get("creator"),
        "last_modified_by": properties.get("lastModifiedBy"),
        "created": properties.get("created"),
        "modified": properties.get("modified"),
        "last_printed": properties.get("lastPrinted"),
    }


def inspect_source(source: dict[str, Any]) -> dict[str, Any]:
    artifact_bytes, workbook_bytes = read_workbook_bytes(source)
    workbook = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=False,
        data_only=False,
        keep_links=True,
    )
    cached_workbook = openpyxl.load_workbook(
        BytesIO(workbook_bytes),
        read_only=False,
        data_only=True,
        keep_links=True,
    )

    defined_names = []
    for defined_name in workbook.defined_names.values():
        defined_names.append(
            {
                "name": defined_name.name,
                "type": defined_name.type,
                "attr_text": defined_name.attr_text,
                "local_sheet_id": defined_name.localSheetId,
                "hidden": defined_name.hidden,
            }
        )

    result = {
        "source_id": source["source_id"],
        "raw_artifact": str(source["artifact"].relative_to(ROOT)).replace("\\", "/"),
        "archive_member": source["member"],
        "version_or_publication": source["version_or_publication"],
        "effective_period": source["effective_period"],
        "retrieved_on": source["retrieved_on"],
        "raw_artifact_bytes": len(artifact_bytes),
        "raw_artifact_sha256": sha256(artifact_bytes),
        "workbook_bytes": len(workbook_bytes),
        "workbook_sha256": sha256(workbook_bytes),
        "core_properties": read_core_properties(workbook_bytes),
        "calculation": {
            "mode": workbook.calculation.calcMode,
            "full_calc_on_load": workbook.calculation.fullCalcOnLoad,
            "force_full_calc": workbook.calculation.forceFullCalc,
        },
        "defined_names": defined_names,
        "sheets": [
            inspect_sheet(
                sheet,
                cached_workbook[sheet.title],
                source["match_terms"],
                source["configured_ranges"].get(sheet.title, ()),
            )
            for sheet in workbook.worksheets
        ],
    }
    workbook.close()
    cached_workbook.close()
    return result


def main() -> int:
    payload = {
        "inspection": {
            "date": "2026-08-03",
            "interpretation_status": "candidate pending source reconciliation",
            "method": "Python openpyxl read-only inspection authorized by user",
            "openpyxl_version": openpyxl.__version__,
            "formula_policy": "formulas captured but not recalculated",
            "numeric_policy": "numeric values serialized lexically; no financial calculation performed",
        },
        "sources": [inspect_source(source) for source in SOURCES],
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    print(OUTPUT)
    return 0


if __name__ == "__main__":
    sys.exit(main())
