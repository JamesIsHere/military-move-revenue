#!/usr/bin/env python3
"""Validate the proposed Item 130 non-monetary fact-model dossier."""

from __future__ import annotations

import copy
import hashlib
import json
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOSSIER_PATH = ROOT / "docs" / "decisions" / "0005-item-130-fact-model-dossier.json"
REGISTRY_PATH = ROOT / "rules" / "registry" / "registry.json"
EXPECTED_ENTITIES = {
    "shipment_article": {"article_id", "shipment_id", "article_kind_observed", "tariff_classification_candidate", "associated_trailer_status", "source_description", "classification_review_status"},
    "article_measurement_observation": {"measurement_id", "article_id", "measurement_kind", "measurement_value", "measurement_unit", "measurement_method", "observed_at", "review_status", "evidence_link_id"},
    "article_condition_observation": {"condition_id", "article_id", "assembled_status", "flat_screen_status", "one_person_hand_carry_status", "standard_carton_transportable_status", "evidence_link_id"},
    "article_service_context": {"service_context_id", "article_id", "shipment_service_code_text", "crating_approval_status", "crating_performed_status", "boto_program_status", "hhg_co_move_agreement_status", "evidence_link_id"},
    "article_handling_event": {"handling_event_id", "article_id", "event_type", "performed_at", "location_role", "performance_status", "shipment_stop_id", "sit_episode_id", "tsp_convenience_status", "evidence_link_id"},
    "combined_handling_occurrence_candidate": {"pairing_candidate_id", "article_id", "loading_event_id", "unloading_event_id", "pairing_status", "pairing_basis", "sit_episode_id", "evidence_link_id"},
    "item_130_preapproval_event": {"approval_event_id", "article_id", "decision_status", "approver_role_observed", "occurred_at", "authorization_reference", "evidence_link_id"},
}
EXPECTED_MISMATCHES = {"GAP-130-LAWNMOWER-ROW", "GAP-130E-SUBTYPE-ROWS", "GAP-130F-BOTO-BOUNDARY", "GAP-130-COMBINED-VS-OD"}
EXPECTED_ROW_DIGEST = "7B124908184726EE6EC38C139CD9FE1C629AC8EE111A1A06632677523A8D6AD5"


class ValidationError(Exception):
    pass


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ValidationError(message)


def load_json(path: Path) -> dict:
    def reject_float(value: str) -> None:
        raise ValidationError(f"{path.name} contains non-exact JSON number {value}")

    with path.open(encoding="utf-8") as handle:
        value = json.load(handle, parse_float=reject_float)
    require(isinstance(value, dict), f"{path.name} must contain an object")
    return value


def workbook_rows(outer_path: Path, sheet: str, first_row: int, last_row: int, columns: int) -> list[list[object]]:
    with ZipFile(outer_path) as outer:
        member = next((name for name in outer.namelist() if name.lower().endswith(".xlsx")), None)
        require(member is not None, f"{outer_path.name} lacks xlsx member")
        workbook = load_workbook(BytesIO(outer.read(member)), read_only=True, data_only=False)
    require(sheet in workbook.sheetnames, f"{outer_path.name} lacks {sheet}")
    worksheet = workbook[sheet]
    return [[worksheet.cell(row, column).value for column in range(1, columns + 1)] for row in range(first_row, last_row + 1)]


def validate(dossier: dict, registry: dict) -> None:
    require(dossier.get("schema_version") == "non-monetary-fact-model-dossier.v1", "dossier schema mismatch")
    require(dossier.get("decision_number") == "0005", "decision number mismatch")
    require(dossier.get("status") == "PROPOSED_OWNER_REVIEW_REQUIRED", "dossier status was advanced without owner review")
    require(dossier.get("prepared_on") == "2026-08-04", "prepared date mismatch")
    require("no field-level or financial interpretation approval" in dossier.get("preparation_authorization", ""), "preparation authority is overstated")
    require(dossier.get("scope") == "DOMESTIC_400NG_ITEM_130_FACT_AND_EVIDENCE_MODEL_ONLY", "scope mismatch")
    require(dossier.get("approval") == {"selected_alternative": None, "approved_by": None, "approved_on": None, "interpretation_decision_id": None}, "approval fields must remain empty")
    require(dossier.get("conflict_ids") == ["CF-0001", "CF-0003"], "conflict scope mismatch")

    conflicts = {record["id"]: record for record in registry["conflict_cases"]}
    require(all(conflicts[conflict_id]["status"] == "open" for conflict_id in dossier["conflict_ids"]), "Item 130 dossier conflict is not open")
    versions = {record["id"]: record for record in registry["source_versions"]}
    source_basis = dossier.get("source_basis")
    require(isinstance(source_basis, list) and len(source_basis) == 3, "source basis must contain three archived records")
    provenance_ids: set[str] = set()
    for source in source_basis:
        provenance_id = source.get("provenance_id")
        require(isinstance(provenance_id, str) and provenance_id.startswith("P-130-") and provenance_id not in provenance_ids, "source provenance id is invalid or duplicated")
        provenance_ids.add(provenance_id)
        version_id = source.get("source_version_id")
        require(version_id in versions and versions[version_id]["source_id"] == source.get("source_id"), "source/version mismatch")
        for field in ("document_version", "effective_period", "locator", "retrieval_date", "interpretation_status"):
            require(isinstance(source.get(field), str) and source[field], f"{provenance_id} lacks {field}")

    internal_basis = dossier.get("internal_basis")
    require(isinstance(internal_basis, list) and len(internal_basis) == 3, "internal basis must contain three records")
    for source in internal_basis:
        provenance_id = source.get("provenance_id")
        require(isinstance(provenance_id, str) and provenance_id.startswith("P-130-") and provenance_id not in provenance_ids, "internal provenance id is invalid or duplicated")
        provenance_ids.add(provenance_id)
        for field in ("document_id", "document_version", "effective_period", "locator", "retrieval_date", "interpretation_status"):
            require(isinstance(source.get(field), str) and source[field], f"{provenance_id} lacks {field}")

    financial = dossier.get("financial_contract")
    require(isinstance(financial, dict) and financial.get("status") == "PROHIBITED", "financial contract is not prohibited")
    for field in ("rate_version_date_fact", "billing_item_contract", "quantity_for_billing", "rate", "expected_amount", "audit_adapter"):
        require(financial.get(field) is None, f"financial field {field} must remain null")

    classifications = dossier.get("tariff_classifications")
    require(isinstance(classifications, list) and [record.get("code") for record in classifications] == [f"130{suffix}" for suffix in "ABCDEFGHIJ"], "tariff classification sequence mismatch")
    require(all(record.get("article_kinds") and record.get("provenance_ids") == ["P-130-TARIFF"] for record in classifications), "tariff classification provenance mismatch")
    require("RIDING_LAWNMOWER_INCLUDING_STAND_ON" in classifications[1]["article_kinds"], "130B lawnmower classification is missing")
    require(len(classifications[4]["article_kinds"]) == 5, "130E watercraft subtype coverage is incomplete")

    entities = dossier.get("proposed_entities")
    require(isinstance(entities, list) and len(entities) == len(EXPECTED_ENTITIES), "entity set is missing or duplicated")
    entity_names: set[str] = set()
    field_count = 0
    for entity in entities:
        name = entity.get("entity")
        require(name in EXPECTED_ENTITIES and name not in entity_names, "entity name is unknown or duplicated")
        entity_names.add(name)
        require(isinstance(entity.get("purpose"), str) and entity["purpose"], f"{name} lacks purpose")
        fields = entity.get("fields")
        require(isinstance(fields, list), f"{name} fields are missing")
        names = [field.get("field") for field in fields]
        require(set(names) == EXPECTED_ENTITIES[name] and len(names) == len(set(names)), f"{name} field contract mismatch")
        field_count += len(fields)
        for field in fields:
            for key in ("logical_type", "cardinality", "evidence_requirement", "interpretation_status"):
                require(isinstance(field.get(key), str) and field[key], f"{name}.{field.get('field')} lacks {key}")
            refs = field.get("provenance_ids")
            require(isinstance(refs, list) and refs and len(refs) == len(set(refs)), f"{name}.{field.get('field')} provenance is invalid")
            require(all(ref in provenance_ids for ref in refs), f"{name}.{field.get('field')} references unknown provenance")

    observation = dossier.get("candidate_item_code_observation")
    require(observation.get("source_provenance_id") == "P-130-ITEM-CODES", "item-code observation provenance mismatch")
    require(observation.get("row_range") == "DOM_400NG!A53:Q118" and observation.get("canonical_row_count") == 66, "item-code observation range mismatch")
    require(observation.get("canonical_rows_sha256") == EXPECTED_ROW_DIGEST, "item-code observation digest mismatch")
    require(observation.get("use_status") == "CANDIDATE_FOR_FUTURE_MAPPING_ONLY_NOT_AN_APPROVED_2026_BILLING_CONTRACT", "item-code observation was promoted")

    mismatches = dossier.get("source_mismatches")
    require(isinstance(mismatches, list) and {record.get("id") for record in mismatches} == EXPECTED_MISMATCHES, "source mismatch set changed")
    for mismatch in mismatches:
        require(mismatch.get("status", "").startswith("OPEN_"), f"{mismatch.get('id')} was silently closed")
        require(mismatch.get("tariff_claim") and mismatch.get("item_code_observation"), f"{mismatch.get('id')} lacks both claims")
        require(mismatch.get("provenance_ids") == ["P-130-TARIFF", "P-130-ITEM-CODES"], f"{mismatch.get('id')} provenance mismatch")

    require([record.get("id") for record in dossier.get("decision_alternatives", [])] == ["A_APPROVE_FACT_MODEL_ONLY", "B_REVISE_FACT_MODEL"], "decision alternatives mismatch")
    tests = dossier.get("mandatory_tests")
    require(isinstance(tests, list) and len(tests) == 18 and len(set(tests)) == 18, "mandatory test contract mismatch")
    required_exclusions = {"EXPECTED_AMOUNT", "RULE_OR_RATE_PACKAGE_PUBLICATION", "ITEM_130_AUDIT_ADAPTER", "REAL_SHIPMENT_DATA"}
    require(required_exclusions.issubset(set(dossier.get("excluded_scope", []))), "required exclusions are missing")
    require(dossier.get("implementation_gate", "").startswith("DO_NOT_REGISTER_AN_ITEM_130"), "implementation stop gate missing")
    require(dossier.get("unresolved_assumptions") == [], "dossier silently carries assumptions")

    item_130_packages = [record for record in registry["rule_packages"] if "ITEM_130" in f"{record.get('id', '')} {record.get('package_code', '')}".upper()]
    item_130_rules = [record for record in registry["rules"] if "ITEM-130" in record.get("id", "").upper()]
    item_130_decisions = [record for record in registry["interpretation_decisions"] if "ITEM 130" in f"{record.get('effective_scope', '')} {record.get('rationale', '')}".upper()]
    require(not item_130_packages and not item_130_rules and not item_130_decisions, "registry contains an unauthorized Item 130 financial contract")

    rows = workbook_rows(ROOT / "sources" / "raw" / "2026" / "item-code-listing-2022-08-12.zip", "DOM_400NG", 53, 118, 17)
    payload = json.dumps(rows, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    require(hashlib.sha256(payload).hexdigest().upper() == EXPECTED_ROW_DIGEST, "archived Item 130 candidate rows changed")
    require(all(row[0] == "A" and row[6] == "EA" and row[10] == 77 and row[16] == "Yes" for row in rows), "candidate row common fields changed")
    require(all(rows[index][4] == ("O" if index % 2 == 0 else "D") for index in range(len(rows))), "candidate origin/destination pairing changed")
    require(not any("lawn" in str(row[5]).lower() for row in rows), "lawnmower gap no longer matches archived rows")
    require([row[5].strip() for row in rows if row[1] == "130E"] == ["Bulky Article: Boats > 14 Ft.", "Bulky Article: Boats > 14 Ft."], "130E candidate rows changed")

    tariff = " ".join((ROOT / "sources" / "derived" / "2026" / "2026-400ng-final.txt").read_text(encoding="utf-8").split())
    required_tariff_text = (
        "Item 130 - Light and/or Bulky Article Classifications",
        "All Bulky Article charges require pre-approval from the Government",
        "One charge for each time a combined loading and unloading service is required",
        "Articles capable of being safely hand-carried by one person",
        "all fractions of a foot will be disregarded",
        "Boat trailers are handled under the BOTO program",
    )
    for text in required_tariff_text:
        require(text in tariff, f"tariff extract lacks {text}")
    library = (ROOT / "sources" / "raw" / "2026" / "ustranscom-dp3-library-2026-08-03.html").read_text(encoding="utf-8", errors="replace")
    require("Item Code Listing (12 Aug 2022)" in library, "library snapshot lacks item-code link")

    validate.field_count = field_count  # type: ignore[attr-defined]


def main() -> int:
    try:
        dossier = load_json(DOSSIER_PATH)
        registry = load_json(REGISTRY_PATH)
        validate(dossier, registry)
        probes = [
            ("status", lambda value: value.__setitem__("status", "APPROVED")),
            ("money", lambda value: value["financial_contract"].__setitem__("expected_amount", "297.78")),
            ("field", lambda value: value["proposed_entities"][0]["fields"].pop()),
            ("mismatch", lambda value: value["source_mismatches"][0].__setitem__("status", "RESOLVED")),
            ("row digest", lambda value: value["candidate_item_code_observation"].__setitem__("canonical_rows_sha256", "0" * 64)),
            ("approval", lambda value: value["approval"].__setitem__("interpretation_decision_id", "INT-0003")),
        ]
        for label, mutate in probes:
            changed = copy.deepcopy(dossier)
            mutate(changed)
            try:
                validate(changed, registry)
            except ValidationError:
                print(f"PASS Item 130 dossier tamper rejected: {label}")
                continue
            raise ValidationError(f"Item 130 dossier tamper accepted: {label}")
        print(f"PASS Item 130 non-monetary dossier: {len(EXPECTED_ENTITIES)} entities, {validate.field_count} fields, 4 source gaps, 18 mandatory tests, and 6 tamper probes")  # type: ignore[attr-defined]
        return 0
    except (OSError, StopIteration, json.JSONDecodeError, KeyError, TypeError, ValidationError) as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
