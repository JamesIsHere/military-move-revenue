#!/usr/bin/env python3
"""Validate the bounded CF-0002 transit/SIT source-research checkpoint."""

from __future__ import annotations

import copy
import hashlib
import json
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RESEARCH_PATH = ROOT / "docs" / "cf-0002-source-research-2026-08-07.json"
REGISTRY_PATH = ROOT / "rules" / "registry" / "registry.json"
ADVISORY_PATH = ROOT / "sources" / "raw" / "2025" / "pcs-jtf-advisory-26-0030-2026-transit-times-2025-12-08.pdf"
ADVISORY_TEXT_PATH = ROOT / "sources" / "derived" / "2025" / "pcs-jtf-advisory-26-0030-2026-transit-times-2025-12-08.txt"
SOLICITATION_PATH = ROOT / "sources" / "raw" / "2025" / "pcs-jtf-advisory-26-0027-2026-rate-filing-event-2025-12-04.pdf"
SOLICITATION_TEXT_PATH = ROOT / "sources" / "derived" / "2025" / "pcs-jtf-advisory-26-0027-2026-rate-filing-event-2025-12-04.txt"
DTR_VJ3_PATH = ROOT / "sources" / "raw" / "2011" / "dtr-part-iv-appendix-v-j-3-intra-country-move-tsp-shipment-management-2011-11-17.pdf"
DTR_VJ3_TEXT_PATH = ROOT / "sources" / "derived" / "2011" / "dtr-part-iv-appendix-v-j-3-sit-cross-reference-2011-11-17.txt"
TARIFF_TEXT_PATH = ROOT / "sources" / "derived" / "2026" / "2026-400ng-final.txt"
TRANSIT_PATH = ROOT / "sources" / "raw" / "2026" / "2026-transit-time-tables.zip"
MILEAGE_SIT_PATH = ROOT / "sources" / "raw" / "2026" / "dps-mileage-transit-time-sit-tool.xlsx"
DTR_TEXT_PATH = ROOT / "sources" / "derived" / "2026" / "dtr-part-iv-chapter-a-402-shipment-management.txt"
EXPECTED_CLAIMS = {
    "CLM-0003",
    "CLM-0004",
    "CLM-0005",
    "CLM-0007",
    "CLM-0008",
    "CLM-0056",
    "CLM-0057",
    "CLM-0058",
    "CLM-0059",
    "CLM-0060",
    "CLM-0061",
    "CLM-0062",
    "CLM-0063",
    "CLM-0064",
}
EXPECTED_FINDINGS = {
    "CF2-TRANSIT-VERSION-SELECTOR": "AUTHORITATIVE_BASIS_IDENTIFIED_RULE_STILL_BLOCKED",
    "CF2-CURRENT-TOOL-IDENTITY": "VERIFIED_BYTE_IDENTICAL_CONFLICT_PERSISTS",
    "CF2-SIT-PERCENTAGE-AUTHORITY": "OPEN_SOLICITATION_IDENTIFIED_PROVISION_NOT_PUBLICLY_LOCATED",
    "CF2-SIT-ROUNDING-AUTHORITY": "OPEN_GOVERNING_ROUNDING_RULE_NOT_LOCATED",
}


class ValidationError(Exception):
    pass


def require(condition: bool, message: str) -> None:
    if not condition:
        raise ValidationError(message)


def load_json(path: Path) -> dict:
    def reject_float(value: str) -> None:
        raise ValidationError(f"{path.name} contains non-exact JSON number {value}")

    with path.open(encoding="utf-8") as handle:
        value = json.load(handle, parse_float=reject_float)
    require(isinstance(value, dict), f"{path.name} must contain an object")
    return value


def validate_raw_sources() -> None:
    advisory = ADVISORY_PATH.read_bytes()
    require(len(advisory) == 119449, "Advisory 26-0030 byte length changed")
    require(
        hashlib.sha256(advisory).hexdigest().upper()
        == "D4161CCA97A96339097503A193FEEF7806D5450E7163DF3C0F231E2C6A37FD6A",
        "Advisory 26-0030 hash changed",
    )
    advisory_text = " ".join(ADVISORY_TEXT_PATH.read_text(encoding="utf-8").split())
    for phrase in (
        "2026 USTC Domestic - International Transit Time Tables",
        "desired pickup date on or after 15 May 2026",
        "No adjustments are being made to other codes of service",
        "2026-USTC Dom-Int Transit Times tables - Effective 15 May 2026.xlsx",
    ):
        require(phrase in advisory_text, f"advisory extract lacks {phrase}")

    solicitation = SOLICITATION_PATH.read_bytes()
    require(len(solicitation) == 711674, "Advisory 26-0027 byte length changed")
    require(
        hashlib.sha256(solicitation).hexdigest().upper()
        == "ACAED24AE6928BBCC3A20768456E1F7A6D780DC06A43B1272F7363D698731986",
        "Advisory 26-0027 hash changed",
    )
    solicitation_text = " ".join(SOLICITATION_TEXT_PATH.read_text(encoding="utf-8").split())
    for phrase in (
        "15 May 2026 through 14 May 2027",
        "Rates Workbench User Guide",
        "The User Guide provides details on functionality",
        "By filing DP3 Rates under this solicitation",
    ):
        require(phrase in solicitation_text, f"solicitation extract lacks {phrase}")

    dtr_vj3 = DTR_VJ3_PATH.read_bytes()
    require(len(dtr_vj3) == 3430521, "DTR Appendix V.J.3 byte length changed")
    require(
        hashlib.sha256(dtr_vj3).hexdigest().upper()
        == "95377E2625849C948170D3BE4BA24CE9E6E3B764C2230A2F4D2FDD3AF757F0D5",
        "DTR Appendix V.J.3 hash changed",
    )
    dtr_vj3_text = " ".join(DTR_VJ3_TEXT_PATH.read_text(encoding="utf-8").split())
    require("percentage (See the International Tender)" in dtr_vj3_text, "historical DTR cross-reference changed")

    tariff_text = " ".join(TARIFF_TEXT_PATH.read_text(encoding="utf-8").split())
    item_29_start = tariff_text.rindex("Item 29 - Tender of Delivery")
    item_29 = tariff_text[item_29_start:tariff_text.index("Item 30", item_29_start)]
    require("If the customer is available to receive property on the TSPs first available delivery date" in item_29, "400NG Item 29 customer/FADD condition changed")
    require("will always be effective on the TSP’s first available delivery date" in item_29, "400NG Item 29 SIT effective-date condition changed")
    require("70%" not in item_29 and "70 percent" not in item_29.lower(), "400NG Item 29 unexpectedly states a 70-percent threshold")

    with ZipFile(TRANSIT_PATH) as archive:
        member = next((name for name in archive.namelist() if name.lower().endswith(".xlsx")), None)
        require(member is not None, "transit archive lacks an XLSX member")
        transit = load_workbook(BytesIO(archive.read(member)), read_only=True, data_only=False)
    require("Appendix L-Domestic" in transit.sheetnames, "transit workbook lacks domestic appendix")
    row = [transit["Appendix L-Domestic"].cell(5, column).value for column in range(1, 7)]
    require(row == ["751-1000", 24, 22, 19, 18, 16], "2026 domestic transit example changed")

    mileage_bytes = MILEAGE_SIT_PATH.read_bytes()
    require(len(mileage_bytes) == 2902386, "mileage/SIT workbook byte length changed")
    require(
        hashlib.sha256(mileage_bytes).hexdigest().upper()
        == "04E62CC2EFAF98FC18536EDBF948BAE6FA3956A9A9FEC86DB1A68BAAA3318DE4",
        "mileage/SIT workbook hash changed",
    )
    mileage = load_workbook(BytesIO(mileage_bytes), read_only=True, data_only=False)
    require(mileage["TT"].sheet_state == "hidden", "mileage/SIT TT sheet is no longer hidden")
    require(
        [mileage["TT"].cell(5, column).value for column in range(1, 7)] == [1000, 14, 12, 10, 9, 8],
        "mileage/SIT conflicting lookup row changed",
    )
    require(mileage["MAIN"]["G10"].value == "*Based on 70% of transit time", "tool percentage label changed")
    require(mileage["WORK"]["I5"].value == "=ROUND(I4*0.7,0)", "tool percentage/rounding formula changed")

    dtr_text = " ".join(DTR_TEXT_PATH.read_text(encoding="utf-8").split())
    require("authorized SIT after a percentage (see solicitation)" in dtr_text, "DTR solicitation cross-reference changed")
    require("applicable percentage of the Government’s transit time" in dtr_text, "DTR percentage-review text changed")


def validate(research: dict, registry: dict) -> None:
    require(research.get("schema_version") == "cf-0002-source-research.v1", "research schema mismatch")
    require(research.get("research_id") == "CF-0002-SOURCE-PASS-2026-08-07", "research id mismatch")
    require(research.get("conducted_on") == "2026-08-07", "research date mismatch")
    require(research.get("status") == "COMPLETED_CONFLICT_REMAINS_OPEN", "research status overstates closure")
    require(
        research.get("scope") == "DOMESTIC_DP3_2026_TRANSIT_VERSION_AND_DIRECT_DELIVERY_SIT_AUTHORITY_ONLY",
        "research scope mismatch",
    )

    versions = {record["id"]: record for record in registry["source_versions"]}
    claims = {record["id"]: record for record in registry["source_claims"]}
    conflicts = {record["id"]: record for record in registry["conflict_cases"]}
    rules = {record["id"]: record for record in registry["rules"]}

    source_basis = research.get("source_basis")
    require(isinstance(source_basis, list) and len(source_basis) == 7, "source basis must contain seven archived records")
    cited_claims: set[str] = set()
    provenance_ids: set[str] = set()
    for source in source_basis:
        provenance_id = source.get("provenance_id")
        require(isinstance(provenance_id, str) and provenance_id.startswith("P-CF2-") and provenance_id not in provenance_ids, "invalid or duplicate provenance id")
        provenance_ids.add(provenance_id)
        version_id = source.get("source_version_id")
        require(version_id in versions and versions[version_id]["source_id"] == source.get("source_id"), f"{provenance_id} source/version mismatch")
        for field in ("document_version", "effective_period", "locator", "retrieval_date", "interpretation_status"):
            require(isinstance(source.get(field), str) and source[field], f"{provenance_id} lacks {field}")
        claim_ids = source.get("source_claim_ids")
        require(isinstance(claim_ids, list) and claim_ids, f"{provenance_id} lacks source claims")
        require(all(claim_id in claims for claim_id in claim_ids), f"{provenance_id} cites an unknown claim")
        cited_claims.update(claim_ids)
    require(cited_claims == EXPECTED_CLAIMS, "research source-claim set changed")

    findings = research.get("findings")
    require(isinstance(findings, list) and len(findings) == 4, "research must contain four findings")
    require({item.get("finding_id"): item.get("status") for item in findings} == EXPECTED_FINDINGS, "finding status set changed")
    for finding in findings:
        require(finding.get("finding") and finding.get("prohibited_inference"), f"{finding.get('finding_id')} lacks finding or stop gate")
        require(all(claim_id in cited_claims for claim_id in finding.get("source_claim_ids", [])), f"{finding.get('finding_id')} cites an unregistered claim")

    conflict = conflicts.get("CF-0002")
    require(conflict is not None and conflict.get("status") == "open", "CF-0002 must remain open")
    expected_rules = {"RULE-DOMESTIC-TRANSIT-TABLE-2026", "RULE-DIRECT-DELIVERY-SIT-DAY-PERCENT"}
    require(set(conflict.get("affected_rule_ids", [])) == expected_rules, "CF-0002 affected rules changed")
    require(rules["RULE-DOMESTIC-TRANSIT-TABLE-2026"].get("effective_date_fact_type") == "desired_pickup_date", "transit selector date fact is not desired pickup date")
    for rule_id in expected_rules:
        rule = rules[rule_id]
        require(rule.get("publication_status") == "draft", f"{rule_id} must remain draft")
        require(rule.get("implementation_status") == "not_implemented", f"{rule_id} must remain unimplemented")
        require(rule.get("blocked_by_conflict_ids") == ["CF-0002"], f"{rule_id} must remain conflict-blocked")

    checkpoints = {item.get("rule_id"): item for item in research.get("rule_checkpoints", [])}
    require(set(checkpoints) == expected_rules, "rule checkpoint set changed")
    require(checkpoints["RULE-DOMESTIC-TRANSIT-TABLE-2026"].get("draft_correction") == "effective_date_fact_type changed from actual_pickup_date to desired_pickup_date", "draft selector correction changed")

    financial = research.get("financial_contract")
    require(isinstance(financial, dict) and financial.get("status") == "PROHIBITED", "financial contract is not prohibited")
    for field in ("rate_version_date_fact", "billing_item_contract", "quantity_for_billing", "rate", "expected_amount", "audit_adapter"):
        require(financial.get(field) is None, f"financial field {field} must remain null")
    external = research.get("external_action")
    require(external.get("status") == "NONE_TAKEN" and external.get("request_sent") is False, "external action status changed")
    validate_raw_sources()


def main() -> int:
    try:
        research = load_json(RESEARCH_PATH)
        registry = load_json(REGISTRY_PATH)
        validate(research, registry)
        probes = [
            ("closure", lambda value: value.__setitem__("status", "RESOLVED")),
            ("finding", lambda value: value["findings"].pop()),
            ("selector", lambda value: value["rule_checkpoints"][0].__setitem__("draft_correction", "actual_pickup_date")),
            ("money", lambda value: value["financial_contract"].__setitem__("expected_amount", "1.00")),
            ("claim", lambda value: value["source_basis"][0]["source_claim_ids"].pop()),
            ("external", lambda value: value["external_action"].__setitem__("request_sent", True)),
        ]
        for label, mutate in probes:
            changed = copy.deepcopy(research)
            mutate(changed)
            try:
                validate(changed, registry)
            except ValidationError:
                print(f"PASS CF-0002 source research tamper rejected: {label}")
                continue
            raise ValidationError(f"CF-0002 source research tamper accepted: {label}")
        print("PASS CF-0002 source research: 7 archived sources, 14 claims, 4 findings, 2 blocked rules, and 6 tamper probes")
        return 0
    except (OSError, StopIteration, json.JSONDecodeError, KeyError, TypeError, ValidationError) as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
