#!/usr/bin/env python3
"""Verify the proposed Item 28B dossier against archived artifacts and gates."""

from __future__ import annotations

import copy
import json
import sys
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DOSSIER = ROOT / "docs" / "decisions" / "0004-item-28b-proposed-dossier.json"
REGISTRY = ROOT / "rules" / "registry" / "registry.json"
ACCEPTED = ROOT / "docs" / "decisions" / "0004-item-28b-scoped-date-and-code.md"


class ValidationError(Exception):
    pass


def require(value: bool, message: str) -> None:
    if not value:
        raise ValidationError(message)


def load(path: Path) -> dict:
    def no_float(value: str) -> None:
        raise ValidationError(f"{path.name} contains JSON float {value}")
    with path.open(encoding="utf-8") as handle:
        value = json.load(handle, parse_float=no_float)
    require(isinstance(value, dict), f"{path.name} must contain an object")
    return value


def workbook_row(outer_path: Path, sheet: str, row: int, count: int) -> list[object]:
    with ZipFile(outer_path) as outer:
        member = next((name for name in outer.namelist() if name.lower().endswith(".xlsx")), None)
        require(member is not None, f"{outer_path.name} lacks xlsx member")
        workbook = load_workbook(BytesIO(outer.read(member)), read_only=True, data_only=False)
    require(sheet in workbook.sheetnames, f"{outer_path.name} lacks {sheet}")
    worksheet = workbook[sheet]
    return [worksheet.cell(row, column).value for column in range(1, count + 1)]


def validate(dossier: dict, registry: dict) -> None:
    require(dossier.get("schema_version") == "interpretation-decision-dossier.v1", "dossier schema mismatch")
    require(dossier.get("decision_number") == "0004", "decision number mismatch")
    require(dossier.get("status") == "PROPOSED_OWNER_APPROVAL_REQUIRED", "dossier was accepted without owner action")
    require(dossier.get("prepared_on") == "2026-08-04", "prepared date mismatch")
    require(dossier.get("approval") == {"selected_alternative": None, "approved_by": None, "approved_on": None, "interpretation_decision_id": None}, "approval fields must remain empty")
    require(dossier.get("conflict_ids") == ["CF-0001", "CF-0003"], "conflict scope mismatch")
    conflicts = {value["id"]: value for value in registry["conflict_cases"]}
    require(all(conflicts[value]["status"] == "open" for value in dossier["conflict_ids"]), "dossier conflict is not open")
    versions = {value["id"]: value for value in registry["source_versions"]}
    require(len(dossier.get("source_basis", [])) == 4, "source basis must contain four records")
    for source in dossier["source_basis"]:
        require(source.get("source_version_id") in versions, "unknown source version")
        require(versions[source["source_version_id"]]["source_id"] == source.get("source_id"), "source/version mismatch")
        for field in ("document_version", "effective_period", "locator", "retrieval_date", "interpretation_status"):
            require(isinstance(source.get(field), str) and source[field], f"source lacks {field}")

    direct = dossier.get("verified_direct_facts")
    require(direct.get("rate") == "198.50" and direct.get("currency") == "USD", "rate contract mismatch")
    require(direct.get("rate_unit") == "USD_per_occurrence" and direct.get("rate_cell") == "Additional Rates!E13", "rate locator/unit mismatch")
    contract = dossier.get("proposed_contract")
    require(contract == {
        "rate_version_date_fact": "actual_pickup_date", "billing_item_code": "28B", "quantity_unit": "EA",
        "rate_basing_reference": "SC", "service_location_role": "ADDITIONAL_DELIVERY", "service_location_code": "AE",
        "approval_screen": "DESTINATION_PPSO", "approval_required": True,
        "calculation": "eligible_completed_extra_delivery_occurrences * 198.50 USD_per_occurrence",
    }, "proposed contract mismatch")
    require([value["id"] for value in dossier.get("alternatives", [])] == ["A_APPROVE_NARROW", "B_DEFER"], "decision alternatives mismatch")
    require(len(dossier.get("mandatory_tests", [])) == 10 and len(set(dossier["mandatory_tests"])) == 10, "mandatory test set mismatch")
    require(dossier.get("implementation_gate", "").startswith("DO_NOT_REGISTER"), "implementation stop gate missing")
    require(dossier.get("unresolved_assumptions") == [], "dossier silently assumes a resolution")

    decisions = {value["id"]: value for value in registry["interpretation_decisions"]}
    rules = {value["id"]: value for value in registry["rules"]}
    packages = {value["id"]: value for value in registry["rule_packages"]}
    decision = decisions.get("INT-0002")
    require(isinstance(decision, dict) and decision.get("decision_status") == "approved", "accepted Item 28B decision is not registered")
    require(decision.get("decided_on") == "2026-08-04" and "explicit agreement" in decision.get("decided_by", ""), "owner approval provenance mismatch")
    require(decision.get("authorized_rule_ids") == ["RULE-ITEM-28B-SCOPED-SOURCE-CONTRACT", "RULE-ITEM-28B-ELIGIBLE-OCCURRENCE", "RULE-ITEM-28B-EXPECTED-CHARGE"], "authorized Item 28B rule scope mismatch")
    item_rules = [rules[value] for value in decision["authorized_rule_ids"]]
    require(all(value.get("publication_status") == "published" and value.get("implementation_status") == "implemented" for value in item_rules), "Item 28B implementation/publication state mismatch")
    package = packages.get("RP-DP3-2026-ITEM-28B-1")
    require(isinstance(package, dict) and package.get("publication_status") == "published", "Item 28B published package gate mismatch")
    accepted_text = ACCEPTED.read_text(encoding="utf-8")
    require("Status: Accepted" in accepted_text and "A_APPROVE_NARROW" in accepted_text and "INT-0002" in accepted_text, "accepted decision record is incomplete")
    rate_row = workbook_row(ROOT / "sources" / "raw" / "2026" / "400ng-baseline-rates.zip", "Additional Rates", 13, 6)
    require(rate_row == [28, "28A,  28B & 28C", None, "Stop off, Diversion, Extra pickups, & Extra Delivery", 198.5, "Per occurrence"], "archived Item 28 rate row changed")
    code_row = workbook_row(ROOT / "sources" / "raw" / "2026" / "item-code-listing-2022-08-12.zip", "DOM_400NG", 24, 17)
    require(code_row == ["A", "28B", "N/A", "dHHG", "D", "Extra Delivery", "EA", None, "SC", "Ex. Del Point (AE)", "AE", None, None, None, None, "Destin PPSO", "Yes"], "archived Item 28B row changed")
    tariff = " ".join((ROOT / "sources" / "derived" / "2026" / "2026-400ng-final.txt").read_text(encoding="utf-8").split())
    for text in ("(28B) - Extra Delivery", "additional deliveries made prior to the final delivery", "each extra pickup or delivery that is performed"):
        require(text in tariff, f"tariff extract lacks {text}")
    library = (ROOT / "sources" / "raw" / "2026" / "ustranscom-dp3-library-2026-08-03.html").read_text(encoding="utf-8", errors="replace")
    require("Item Code Listing (12 Aug 2022)" in library, "library snapshot lacks item-code link")


def main() -> int:
    try:
        dossier, registry = load(DOSSIER), load(REGISTRY)
        validate(dossier, registry)
        probes = [("status", "ACCEPTED"), ("rate", "198.51"), ("date", "original_requested_pickup_date"), ("approval", "INT-0002")]
        for label, value in probes:
            changed = copy.deepcopy(dossier)
            if label == "status": changed["status"] = value
            elif label == "rate": changed["verified_direct_facts"]["rate"] = value
            elif label == "date": changed["proposed_contract"]["rate_version_date_fact"] = value
            else: changed["approval"]["interpretation_decision_id"] = value
            try: validate(changed, registry)
            except ValidationError:
                print(f"PASS Item 28B dossier tamper rejected: {label}")
                continue
            raise ValidationError(f"accepted dossier tamper: {label}")
        print("PASS preserved Item 28B proposal, accepted INT-0002 published-package gate, 4 archived sources, 10 mandatory tests, and 4 tamper probes")
        return 0
    except (OSError, StopIteration, json.JSONDecodeError, KeyError, TypeError, ValidationError) as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
